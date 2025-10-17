import os
import re
import tempfile
from typing import Optional, Dict, Any, List
import requests
from openpyxl.cell import Cell

from config import LESSON_TIMES, GROUP_CODES


class ExcelParser:
    def __init__(self):
        self.temp_files: List[str] = []

    def get_group_schedule(self, excel_content: str, group_name: str) -> Optional[Dict[str, Any]]:

        from openpyxl import load_workbook

        try:
            total_lessons = 0
            distant_lessons = 0
            self_study_lessons = 0
            normal_lessons = 0
            current_day = "Понедельник"
            excel_path = self.download_excel(excel_content)
            if not excel_path:
                return None

            wb = load_workbook(excel_path)
            ws = wb.active

            # Определяем строку с группами в зависимости от файла
            filename = excel_content.lower()
            if "2-3" in filename:
                group_row = 7  # для файла 2-3 курсы
            else:
                group_row = 6  # для остальных файлов

            print(f"🎯 Ищу группу в строке {group_row}")

            # Ищем группу в правильной строке
            group_col = None
            for col in range(4, ws.max_column + 1):
                cell_value = ws.cell(row=group_row, column=col).value
                if cell_value and group_name.upper() in str(cell_value).upper():
                    group_col = col
                    print(f"✅ Найдена группа '{group_name}' в колонке {col}")
                    break

            if not group_col:
                print(f"❌ Группа '{group_name}' не найдена в файле!")
                return None

            schedule = {}

            # Парсим начиная со следующей строки после заголовка
            start_row = group_row + 1
            print(f"🎯 Начинаю парсинг с строки {start_row}")

            for row in range(start_row, ws.max_row + 1):
                day_cell = ws.cell(row=row, column=1)  # ячейка дня
                time_cell = ws.cell(row=row, column=2)  # ячейка времени
                lesson_num_cell = ws.cell(row=row, column=3)  # ячейка номера пары
                lesson_cell = ws.cell(row=row, column=group_col)  # ячейка предмета

                # Обновляем текущий день
                if day_cell.value and str(day_cell.value).strip():
                    current_day = str(day_cell.value).strip().split()[0]

                if not lesson_num_cell.value:
                    continue

                try:
                    lesson_num = int(lesson_num_cell.value)
                except (ValueError, TypeError):
                    continue

                if not lesson_cell.value or not str(lesson_cell.value).strip():
                    continue

                color_type = self.get_cell_color_type(lesson_cell)

                total_lessons += 1
                if color_type == "distant":
                    distant_lessons += 1
                elif color_type == "self_study":
                    self_study_lessons += 1
                else:
                    normal_lessons += 1

                lesson_text = str(lesson_cell.value).strip()
                parsed = self.parse_lesson_text(lesson_text)

                lesson_time = self.get_lesson_time(lesson_num)

                if color_type == "distant":
                    subject_text = f"💻 {parsed['subject']} (дистант)"
                elif color_type == "self_study":
                    subject_text = f"📚 {parsed['subject']} (самостоятельная)"
                else:
                    subject_text = parsed['subject']

                schedule[lesson_num] = {
                    "day": current_day,
                    "time": lesson_time,
                    "subject": subject_text,
                    "teacher": parsed["teacher"],
                    "room": parsed["room"],
                    "color_type": color_type,
                    "subgroup": parsed.get("subgroup", "")
                }

            wb.close()

            stats_data = {
                "total": total_lessons,
                "distant": distant_lessons,
                "self_study": self_study_lessons,
                "normal": normal_lessons
            }

            print(f"📊 Расписание собрано: {len(schedule)} пар")

            return {
                "schedule": schedule,
                "stats": stats_data
            }

        except Exception as e:
            print(f"❌ Ошибка парсинга расписания: {e}")
            return None
        finally:
            if isinstance(excel_content, str) and excel_content.startswith('http'):
                self.cleanup_temp_files()

    def download_excel(self, url: str) -> Optional[str]:
        try:
            if url.startswith('http'):
                response = requests.get(url, timeout=30)
                response.raise_for_status()

                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file.write(response.content)
                temp_file.close()

                self.temp_files.append(temp_file.name)
                return temp_file.name
            else:
                # Для локальных файлов - создаем копию
                if os.path.exists(url):
                    import shutil
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    shutil.copy2(url, temp_file.name)
                    self.temp_files.append(temp_file.name)
                    return temp_file.name
                return None
        except Exception as e:
            print(f"❌ Ошибка загрузки: {e}")
            return None

    def cleanup_temp_files(self) -> None:
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                print(f"⚠️ Не удалось удалить временный файл: {e}")
        self.temp_files.clear()

    @staticmethod
    def get_cell_color_type(cell: Cell) -> str:
        try:
            if cell.fill.start_color and cell.fill.start_color.index == 9:
                return "distant"  # Единственная цветная ячейка
            return "normal"
        except (AttributeError, TypeError, ValueError):
            return "normal"

    @staticmethod
    def is_valid_group_name(group_name: str) -> bool:
        """Проверяет, является ли текст названием группы"""
        text = str(group_name).strip()

        # Пропускаем пустые
        if not text:
            return False

        # Пропускаем очевидные не-группы
        if any(word in text.lower() for word in ['расписание', 'занятий', 'семестр', 'учебный', 'год', '№']):
            return False

        # Группы обычно содержат буквы и цифры, короткие
        if len(text) > 15:
            return False

        # Должны быть и буквы и цифры
        has_letters = any(c.isalpha() for c in text)
        has_digits = any(c.isdigit() for c in text)

        return has_letters and has_digits

    @staticmethod
    def is_group_in_course(group_name: str, course: str) -> bool:
        """Проверяет, принадлежит ли группа к курсу по буквенному коду"""
        group_codes = GROUP_CODES.get(course, [])
        group_name_str = str(group_name).upper().strip()

        # Ищем любой из кодов групп в названии
        for code in group_codes:
            if code.upper() in group_name_str:
                return True
        return False

    def find_groups_in_excel(self, excel_content, course_name):
        """
        Ищет названия групп в Excel в зависимости от файла:
        - 1 курсы / 1-2 курсы / 3-4 курсы → строка 6 (индекс 5)
        - 2-3 курсы → строка 7 (индекс 6)
        """

        try:
            import pandas as pd
            from io import BytesIO
            import requests

            # Загружаем файл: URL или локальный путь
            if isinstance(excel_content, str) and excel_content.lower().startswith("http"):
                file_data = BytesIO(requests.get(excel_content).content)
                df = pd.read_excel(file_data, header=None)
            else:
                df = pd.read_excel(excel_content, header=None)

            # Определяем строку с группами
            file = str(excel_content).lower()
            if "2-3" in file:
                row_index = 6  # 7 строка
            else:
                row_index = 5  # 6 строка

            group_row = df.iloc[row_index]
            groups = []

            for value in group_row[2:]:  # начиная с колонки C
                if isinstance(value, str) and value.strip():
                    groups.append(value.strip())

                    # Убираем лишние значения типа "№" или None
                    groups = [g for g in groups if isinstance(g, str) and len(g) > 2 and g[0].isalnum()]

            return groups

        except Exception as e:
            print(f"Ошибка при поиске групп: {e}")
            return []

    @staticmethod
    def parse_lesson_text(lesson_text: str) -> Dict[str, str]:
        text = str(lesson_text).strip()

        if not text:
            return {"subject": "❌ Нет пары", "teacher": "", "room": "", "subgroup": ""}

        text = ' '.join(text.split())
        print(f"🔍 Начало парсинга: '{text}'")

        # 1. Подгруппа
        subgroup = ""
        subgroup_patterns = [r'(\d\s?и\s?\d\s?[п]?од?гр?)', r'(\d\s?[п]?од?гр?)']
        for pattern in subgroup_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                subgroup = match.group(1)
                text = re.sub(pattern, '', text, flags=re.IGNORECASE).strip()
                break

        print(f"📌 Подгруппа: '{subgroup}'")

        # 2. Ищем числа (потенциальные аудитории)
        raw_rooms = re.findall(r'\b(\d{2,4}[A-ZА-Я]?)\b', text)
        print(f"🚪 Найденные числа: {raw_rooms}")

        room = ""
        if raw_rooms:
            room = raw_rooms[-1]  # последняя — аудитория
            text = re.sub(r'\b' + re.escape(room) + r'\b', ' ', text).strip()

            other_codes = raw_rooms[:-1]
            if other_codes:
                text = " ".join(other_codes) + " " + text

        print(f"🚪 Итоговая аудитория: '{room}'")
        print(f"📎 Текст после аудиторий: '{text}'")

        # 3. Известные преподаватели
        known_teachers = [
            'Шпейт', 'Таран', 'Морозова', 'Соколова', 'Олешкевич', 'Догадин',
            'Денисов', 'Зыкова', 'Лобанов', 'Коротков', 'Бухатиева', 'Коврижных',
            'Гоголева', 'Губич', "Банина", "Тухланова", "Артынгова", "Криницин",
            "Криницина", "Воронова", "Дражник", "Кудина", "Киселев", "Чичигина",
            "Земцов", "Усатов", "Колганов", "Сидорова", "Волковинская", "Кочергина",
            "Суслина", "Белинская", "Олешкевич", "Кувалдин", "Владимиров", "Вяземская",
            "Ващенко", "Ромашина", "Гаврилец", "Лынкин", "Бузаев", "Щербаченя", "Костюченко",
            "Белошапкин", "Еремина", "Трифонова", "Вяткина", "Комаристов", "Гусева", "Баженова",
            "Цыганкова", "Окладников", "Чиркова", "Котыхова", "Тварадзе", "Егоров", "Максимова",
            "Селюн", "Падалко", "Торосян", "Стрижаков", "Невина"
        ]

        teachers = []
        for word in text.split():
            clean = re.sub(r'[^А-Яа-я]', '', word)
            if clean in known_teachers:
                teachers.append(clean)
                text = text.replace(word, ' ', 1).strip()

        print(f"👨‍🏫 Преподаватели: {teachers}")

        subject = ' '.join(text.split()).strip()
        subject = re.sub(r'^[,\s\-–—()]+|[,\s\-–—()]+$', '', subject)

        print(f"🎯 Итог subject='{subject}', teacher='{', '.join(teachers)}', room='{room}', subgroup='{subgroup}'")

        return {
            "subject": subject or "?",
            "teacher": ", ".join(teachers),
            "room": room,
            "subgroup": subgroup
        }

    @staticmethod
    def get_lesson_time(lesson_num: int) -> str:
        return LESSON_TIMES.get(lesson_num, "?")
